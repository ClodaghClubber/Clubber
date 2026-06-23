from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime, re

def to_iso(d):
    day, mon, year = d.split('/')
    return f"{year}-{mon}-{day}"

def day_name(iso):
    y, m, d = map(int, iso.split('-'))
    return datetime.date(y, m, d).strftime('%A')

def to12h(t):
    m = re.match(r'(\d{1,2})-(\d{2})(am|pm)', t.strip(), re.I)
    h, mi, ap = m.group(1), m.group(2), m.group(3).upper()
    return f"{int(h)}:{mi} {ap}"

rows = []

def add(comp, group, data):
    for teamA, teamB, venue, date, time, rnd in data:
        iso = to_iso(date)
        rows.append(['Kildare', teamA, teamB, iso, day_name(iso), to12h(time), venue, f"{comp} Group {group}", rnd, 'Proposed'])

add('Intermediate Football Championship', 'A', [
    ('Leixlip', 'Suncroft', 'Conneff Park Clane', '08/08/2026', '5-00pm', 'Round 1'),
    ('Confey', 'Straffan', "Cedral St Conleth's", '08/08/2026', '5-15pm', 'Round 1'),
    ('Confey', 'Leixlip', 'Conneff Park Clane', '22/08/2026', '3-30pm', 'Round 2'),
    ('Straffan', 'Suncroft', 'Manguard Park Pitch 2', '23/08/2026', '1-30pm', 'Round 2'),
    ('Confey', 'Suncroft', 'Manguard Park Pitch 1', '04/09/2026', '8-00pm', 'Round 3'),
    ('Leixlip', 'Straffan', 'Manguard Park Pitch 2', '04/09/2026', '8-00pm', 'Round 3'),
])
add('Intermediate Football Championship', 'B', [
    ('Kilcullen', 'Monasterevan', "Cedral St Conleth's", '08/08/2026', '3-30pm', 'Round 1'),
    ('Round Towers', "St. Laurence's", "Cedral St Conleth's", '09/08/2026', '3-45pm', 'Round 1'),
    ('Kilcullen', "St. Laurence's", 'Manguard Park Pitch 2', '22/08/2026', '5-00pm', 'Round 2'),
    ('Monasterevan', 'Round Towers', "Cedral St Conleth's", '23/08/2026', '2-00pm', 'Round 2'),
    ('Kilcullen', 'Round Towers', 'Raheens GAA', '05/09/2026', '6-00pm', 'Round 3'),
    ('Monasterevan', "St. Laurence's", 'Manguard Park Pitch 1', '05/09/2026', '6-00pm', 'Round 3'),
])
add('Intermediate Football Championship', 'C', [
    ('Ballyteague', 'Milltown', "Cedral St Conleth's", '07/08/2026', '7-45pm', 'Round 1'),
    ('Grangenolvin', 'Nurney', 'Manguard Park Pitch 1', '08/08/2026', '5-00pm', 'Round 1'),
    ('Ballyteague', 'Nurney', 'Manguard Park Pitch 2', '22/08/2026', '3-30pm', 'Round 2'),
    ('Grangenolvin', 'Milltown', 'Manguard Park Pitch 1', '23/08/2026', '6-00pm', 'Round 2'),
    ('Ballyteague', 'Grangenolvin', 'Kilcullen', '06/09/2026', '2-00pm', 'Round 3'),
    ('Milltown', 'Nurney', 'Manguard Park Pitch 2', '06/09/2026', '2-00pm', 'Round 3'),
])
add('Intermediate Football Championship', 'D', [
    ('Castledermot', 'Two Mile House', 'Manguard Park Pitch 1', '08/08/2026', '6-30pm', 'Round 1'),
    ('Ellistown', 'Rathangan', "Cedral St Conleth's", '09/08/2026', '2-00pm', 'Round 1'),
    ('Castledermot', 'Rathangan', 'Manguard Park Pitch 1', '23/08/2026', '3-00pm', 'Round 2'),
    ('Ellistown', 'Two Mile House', 'Manguard Park Pitch 2', '23/08/2026', '4-30pm', 'Round 2'),
    ('Ellistown', 'Castledermot', "St Laurence's", '06/09/2026', '3-30pm', 'Round 3'),
    ('Rathangan', 'Two Mile House', 'Round Towers GFC', '06/09/2026', '3-30pm', 'Round 3'),
])

add('Senior Football Championship', 'A', [
    ('Clane', 'Johnstownbridge', 'Manguard Park Pitch 1', '21/08/2026', '8-00pm', 'Round 1'),
    ('Allenwood', 'Celbridge', 'Conneff Park Clane', '22/08/2026', '5-00pm', 'Round 1'),
    ('Allenwood', 'Clane', "Cedral St Conleth's", '06/09/2026', '3-45pm', 'Round 2'),
    ('Celbridge', 'Johnstownbridge', 'Manguard Park Pitch 1', '06/09/2026', '5-00pm', 'Round 2'),
    ('Allenwood', 'Johnstownbridge', 'Manguard Park Pitch 1', '18/09/2026', '8-00pm', 'Round 3'),
    ('Celbridge', 'Clane', "Cedral St Conleth's", '18/09/2026', '8-00pm', 'Round 3'),
])
add('Senior Football Championship', 'B', [
    ('Carbury', 'Sarsfields', 'Manguard Park Pitch 1', '22/08/2026', '2-00pm', 'Round 1'),
    ('Caragh', 'Maynooth', "Cedral St Conleth's", '22/08/2026', '3-30pm', 'Round 1'),
    ('Caragh', 'Sarsfields', "Cedral St Conleth's", '04/09/2026', '7-45pm', 'Round 2'),
    ('Carbury', 'Maynooth', "Cedral St Conleth's", '05/09/2026', '5-15pm', 'Round 2'),
    ('Maynooth', 'Sarsfields', "Cedral St Conleth's", '19/09/2026', '3-00pm', 'Round 3'),
    ('Caragh', 'Carbury', 'Manguard Park Pitch 1', '19/09/2026', '3-00pm', 'Round 3'),
])
add('Senior Football Championship', 'C', [
    ('Raheens', 'Sallins', "Cedral St Conleth's", '22/08/2026', '5-00pm', 'Round 1'),
    ('Athy', 'Kilcock', "Cedral St Conleth's", '23/08/2026', '3-45pm', 'Round 1'),
    ('Athy', 'Sallins', "Cedral St Conleth's", '06/09/2026', '2-00pm', 'Round 2'),
    ('Kilcock', 'Raheens', 'Manguard Park Pitch 1', '06/09/2026', '3-30pm', 'Round 2'),
    ('Kilcock', 'Sallins', 'Manguard Park Pitch 1', '20/09/2026', '5-00pm', 'Round 3'),
    ('Athy', 'Raheens', "Cedral St Conleth's", '20/09/2026', '5-00pm', 'Round 3'),
])
add('Senior Football Championship', 'D', [
    ('Moorefield', 'Naas', "Cedral St Conleth's", '21/08/2026', '7-45pm', 'Round 1'),
    ('Clogherinkoe', 'Eadestown', 'Manguard Park Pitch 1', '22/08/2026', '6-30pm', 'Round 1'),
    ('Clogherinkoe', 'Moorefield', 'Manguard Park Pitch 1', '05/09/2026', '3-00pm', 'Round 2'),
    ('Eadestown', 'Naas', "Cedral St Conleth's", '05/09/2026', '3-30pm', 'Round 2'),
    ('Eadestown', 'Moorefield', "Cedral St Conleth's", '19/09/2026', '4-45pm', 'Round 3'),
    ('Clogherinkoe', 'Naas', 'Manguard Park Pitch 1', '19/09/2026', '4-45pm', 'Round 3'),
])

add('Junior Football Championship', 'A', [
    ('Robertstown', "St Kevin's", 'Conneff Park Clane', '08/08/2026', '3-30pm', 'Round 1'),
    ('Rathcoffey', 'Rheban', 'Manguard Park Pitch 1', '09/08/2026', '1-30pm', 'Round 1'),
    ('Rheban', "St Kevin's", 'Round Towers GFC', '22/08/2026', '5-00pm', 'Round 2'),
    ('Rathcoffey', 'Robertstown', 'Conneff Park Clane', '23/08/2026', '3-30pm', 'Round 2'),
    ('Rathcoffey', "St Kevin's", 'Conneff Park Clane', '05/09/2026', '4-30pm', 'Round 3'),
    ('Rheban', 'Robertstown', 'Manguard Park Pitch 2', '05/09/2026', '4-30pm', 'Round 3'),
])
add('Junior Football Championship', 'B', [
    ('Cappagh', 'Castlemitchell', 'Manguard Park Pitch 1', '08/08/2026', '3-30pm', 'Round 1'),
    ('Ballykelly', 'Kildangan', 'Manguard Park Pitch 1', '09/08/2026', '3-00pm', 'Round 1'),
    ('Ballykelly', 'Cappagh', 'Conneff Park Clane', '23/08/2026', '2-00pm', 'Round 2'),
    ('Castlemitchell', 'Kildangan', 'Kilcullen', '23/08/2026', '2-00pm', 'Round 2'),
    ('Ballykelly', 'Castlemitchell', 'Round Towers GFC', '05/09/2026', '3-00pm', 'Round 3'),
    ('Cappagh', 'Kildangan', 'Conneff Park Clane', '05/09/2026', '3-00pm', 'Round 3'),
])
add('Junior Football Championship', 'C', [
    ('Ardclough', 'Athgarvan', 'Manguard Park Pitch 1', '07/08/2026', '8-00pm', 'Round 1'),
    ('Ballymore Eustace', 'Kill', 'Manguard Park Pitch 1', '09/08/2026', '4-30pm', 'Round 1'),
    ('Ardclough', 'Ballymore Eustace', 'Manguard Park Pitch 2', '21/08/2026', '7-45pm', 'Round 2'),
    ('Athgarvan', 'Kill', 'Raheens GAA', '23/08/2026', '2-00pm', 'Round 2'),
    ('Athgarvan', 'Ballymore Eustace', 'Manguard Park Pitch 1', '03/09/2026', '8-00pm', 'Round 3'),
    ('Ardclough', 'Kill', 'Manguard Park Pitch 2', '03/09/2026', '8-00pm', 'Round 3'),
])

print('total rows:', len(rows))

wb = Workbook()
ws = wb.active
ws.title = 'Kildare Fixtures'
headers = ['County', 'Team A', 'Team B', 'Date', 'Day', 'Time', 'Venue', 'Competition', 'Round', 'Status']
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E78')
    cell.alignment = Alignment(horizontal='center')

for r in rows:
    ws.append(r)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name='Arial')

widths = [10, 16, 16, 12, 11, 9, 22, 32, 9, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

ws.freeze_panes = 'A2'
wb.save('kildare-fixtures.xlsx')
print('saved')
